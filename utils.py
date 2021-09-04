from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
from PyPDF2 import PdfFileWriter, PdfFileReader
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4


def feeder_name(txt):  # function to strip feeder name from 'TEXT'
    txt_u = txt.upper()
    start_index = txt_u.find('_CB ') + 4
    end_index = txt_u.find(' STTS')
    return txt[start_index: end_index]


def status(txt):
    txt_u = txt.upper()
    if 'OPEN' in txt_u:
        return 0
    elif 'CLOSE' in txt_u:
        return 1
    elif 'BETWEEN' in txt_u:
        return 2
    else:
        return 3


def operator_id(txt):
    if 'operator' in txt.lower():
        return txt[txt.lower().find('operator'):]
    elif 'scada' in txt.lower():
        return 'OP ID N/A'
    else:
        return 'manually'


def decorator(openpyxl_sheet):
    ws = openpyxl_sheet

    # adjusting column widths
    # enumerate index 1st column as 1
    for i, col in enumerate(ws.iter_cols(values_only=True), 1):
        # print(i)
        max_len = 0
        for cell in col:
            if cell is not None:
                if len(str(cell)) > max_len:
                    max_len = len(str(cell))
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = max_len * .5 + 6
        print(f'width of col {get_column_letter(i)} is {max_len + 3}')

    # adjusting borders
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    return ws


def pdf_decorator(existing, modified):
    packet = io.BytesIO()
    # create a new PDF with Reportlab
    can = canvas.Canvas(packet, pagesize=A4)
    can.setFont('Helvetica', 8)
    can.drawString(0, 1, "NDT V1.0 ©2021")
    can.setFont('Helvetica', 10)
    can.drawString(120, 30, '       Prepared By')
    can.drawString(120, 20, 'Sharif Shamsul Alam')
    can.drawString(120, 10, 'AE, IMD, LDC, PGCB')
    can.save()

    # move to the beginning of the StringIO buffer
    packet.seek(0)
    new_pdf = PdfFileReader(packet)
    # read your existing PDF
    existing_pdf = PdfFileReader(open(existing, "rb"))
    output = PdfFileWriter()
    # add the "watermark" (which is the new pdf) on the existing page
    # for i in range(existing_pdf.getNumPages()):
    n_page = existing_pdf.getNumPages()
    for i in range(n_page):
        page = existing_pdf.getPage(i)
        page.mergePage(new_pdf.getPage(0))
        output.addPage(page)
    # finally, write "output" to a real file
    outputStream = open(modified, "wb")
    output.write(outputStream)
    outputStream.close()
