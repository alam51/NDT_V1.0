import openpyxl
import pandas as pd
import re

import win32com.client
from openpyxl.styles import numbers, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas import ExcelWriter
import utils


def scada_reporter():
    # file path input
    file_path = re.escape(input('File path:   '))
    # file_path = r'F:\\'
    file_name = input('File name:  ')
    # file_name = 'export (21)'
    # file_type = input('File type:  ')
    file_type = '.csv'
    absolute_file_path = file_path + '\\' + file_name + file_type

    # file import. 'Dayfirst' should be input carefully
    df = pd.read_csv(absolute_file_path, index_col=None, parse_dates=[0], infer_datetime_format=True,
                     low_memory=False, dayfirst=False)
    df.dropna(axis=1, thresh=int(len(df.iloc[:, 0]) * .9), inplace=True)

    # Separating same timestamps of multiple operations to different time in milli second intervals
    n_rows = len(df.loc[:, 'TIME'])
    i = 0
    while i < n_rows:
        for k in range(1, n_rows - i):
            if df.loc[i, 'TIME'] == df.loc[i + k, 'TIME']:
                df.loc[i + k, 'TIME'] += pd.to_timedelta(str(k) + 'ms')
            else:
                break
        i = i + k

    # indexing by time
    df1 = df.set_index('TIME')
    print(f'Data available from {df1.index[0]} to {df1.index[-1]}')
    t_strt = input('Start Time [yy-mm-dd]: ')
    t_strt = str(pd.to_datetime(t_strt, dayfirst=False, yearfirst=True)) if t_strt != ''\
        else str(df1.index[0])[:10]
    t_end = input('End Time [yy-mm-dd]: ')
    t_end = str(pd.to_datetime(t_end, dayfirst=False, yearfirst=True)) if t_end != ''\
        else str(df1.index[-1])[:10]

    df = df1[t_strt:t_end if pd.to_datetime(t_end) < df1.index[-1] else None]
    # main processing
    SCADA_open_times = []
    SCADA_opened_close_times = []
    SCADA_opened_SS_list = []
    SCADA_opened_feeder_list = []
    close_time_found_list = []
    SCADA_opened_by_list = []
    SCADA_closed_by_list = []
    count = 0
    print(df.index)
    for i in df.index:
        # assume that close time is not found (status = 2)
        close_time_found = 3
        text = df.loc[i, 'TEXT']
        if 'BY' in text.upper() and utils.status(text) == 0:
            SCADA_opened_feeder_list.append(utils.feeder_name(text))
            SCADA_opened_SS_list.append(df.loc[i, 'LOCATION'])
            SCADA_open_times.append(i)
            SCADA_opened_by_list.append(utils.operator_id(text))
            for j in df.index[count + 1:]:
                text_1 = df.loc[j, 'TEXT']
                if utils.status(text_1) == 1 and utils.feeder_name(text_1) == utils.feeder_name(text):
                    SCADA_opened_close_times.append(j)
                    SCADA_closed_by_list.append(utils.operator_id(text_1))
                    close_time_found = 1
                    # close time found (status = 1)
                    break

            # if  close time not found, try for 'BETWEEN' status (status = 2)
            if close_time_found != 1:
                for j in df.index[count + 1:]:
                    text_1 = df.loc[j, 'TEXT']
                    if utils.status(text_1) == 2 and \
                            utils.feeder_name(text_1) == utils.feeder_name(text):
                        SCADA_opened_close_times.append(j)
                        SCADA_closed_by_list.append(utils.operator_id(text_1))
                        close_time_found = 2
                        # close time found
                        break

            # if  not found even in 'BETWEEN' status (status = 2), declare that not found (status = 3)
            if close_time_found != 1 and close_time_found != 2:
                close_time_found = 3
                # add a blank in time
                SCADA_opened_close_times.append(pd.NaT)
                SCADA_closed_by_list.append(pd.NA)
            close_time_found_list.append(close_time_found)
        count += 1

    # print(count)
    # print(df)
    t_open = pd.to_datetime(SCADA_open_times)
    t_close = pd.to_datetime(SCADA_opened_close_times)
    duration = t_close - t_open
    duration_str = []

    d = {'SS': SCADA_opened_SS_list,
         'Feeder': SCADA_opened_feeder_list,
         'Open Time': t_open,
         'Opened by': SCADA_opened_by_list,
         'Close Time': t_close,
         'Closed By': SCADA_closed_by_list,
         'Duration': duration.round('min'),
         'Finding_Status': close_time_found_list
         }
    df2 = pd.DataFrame(data=d, index=None)

    # wb = openpyxl.workbook
    output_path = re.escape(input('Output_folder: ')) + '\\' + 'SCADA_rep' + t_strt[:10] + \
                  'to' + t_end[:10] + '.xlsx'
    writer = pd.ExcelWriter(output_path,
                            date_format='YYYY-MM-DD',
                            datetime_format='YYYY-MM-DD HH:MM:SS', engine='openpyxl')
    df2.to_excel(writer, sheet_name='1')
    wb = writer.book
    ws = writer.sheets['1']
    # ws.delete_cols(1)
    # format1 = workbook.add_format({'num_format': 'h:mm:ss'})
    # worksheet.set_column('F:F', 36, None)
    col = ws['H']
    for i in col:
        i.number_format = numbers.FORMAT_DATE_TIMEDELTA
        # print(i)

    writer.sheets['1'] = utils.decorator(ws)
    ws.insert_rows(1, 3)
    font = Font(name='Calibri',
                size=13,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
    ws['A1'].value = 'National Load Dispatch Center'
    ws['A1'].font = font
    ws.merge_cells('A1:I1')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'].value = '33kV Feeders SCADA Operations Report'
    ws.merge_cells('A2:I2')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A3'].value = t_strt + ' to ' + t_end
    ws.merge_cells('A3:I3')
    ws['A3'].alignment = Alignment(horizontal='center')
    # ws.set_printer_settings(0, ORIENTATION_PORTRAIT)
    ws.page_setup.paperHeight = '210mm'
    ws.page_setup.paperWidth = '297mm'
    ws.print_title_rows = '1:4'  # First four rows are printed in each page
    writer.save()
    print('Successfully saved at : ' + output_path)
    # writer.close()

    # *****pdf creation******
    o = win32com.client.Dispatch("Excel.Application")
    o.Visible = False
    wb_path = output_path
    wb = o.Workbooks.Open(wb_path)
    ws_index_list = [1]  # start from 1
    # path_to_pdf = r'C:\Users\hE\Desktop\sample.pdf'
    path_to_pdf = wb_path.replace('.xlsx', '.pdf')
    print_area = 'A1:I' + str(len(df2.index) + 3)
    for index in ws_index_list:
        # off-by-one so the user can start numbering the worksheets at 1
        ws = wb.Worksheets[index - 1]
        ws.PageSetup.Zoom = False
        ws.PageSetup.FitToPagesTall = 1
        ws.PageSetup.FitToPagesWide = 1
        ws.PageSetup.PrintArea = print_area
    wb.WorkSheets(ws_index_list).Select()
    wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
    utils.pdf_decorator(path_to_pdf, path_to_pdf.replace('.pdf', '_marked.pdf'))
